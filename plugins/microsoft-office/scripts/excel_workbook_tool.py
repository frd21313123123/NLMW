#!/usr/bin/env python3
"""Inspect and lightly edit Microsoft Excel workbooks."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Iterable


def require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl. Install it or load the bundled workspace dependencies."
        ) from exc
    return openpyxl


def load_workbook(path: Path, data_only: bool = False, read_only: bool = False):
    openpyxl = require_openpyxl()
    if not path.exists():
        raise SystemExit(f"File not found: {path}")
    keep_vba = path.suffix.lower() == ".xlsm"
    return openpyxl.load_workbook(
        filename=str(path),
        data_only=data_only,
        read_only=read_only,
        keep_vba=keep_vba,
    )


def cell_value_for_json(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def sheet_by_name(workbook, name: str):
    if name not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {name}. Available sheets: {', '.join(workbook.sheetnames)}")
    return workbook[name]


def iter_used_cells(worksheet) -> Iterable[object]:
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                yield cell


def inspect_workbook(args) -> int:
    path = Path(args.workbook)
    workbook = load_workbook(path, data_only=False)
    sheets = []

    for worksheet in workbook.worksheets:
        formula_count = 0
        non_empty_count = 0
        for cell in iter_used_cells(worksheet):
            non_empty_count += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1

        sheets.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "dimensions": worksheet.calculate_dimension(),
                "non_empty_cells": non_empty_count,
                "formula_cells": formula_count,
                "tables": len(getattr(worksheet, "tables", {})),
                "charts": len(getattr(worksheet, "_charts", [])),
                "merged_ranges": [str(item) for item in list(worksheet.merged_cells.ranges)[:50]],
                "freeze_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
            }
        )

    result = {
        "file": str(path),
        "workbook_type": path.suffix.lower(),
        "sheet_count": len(workbook.sheetnames),
        "sheets": sheets,
        "defined_names": [name.name for name in workbook.defined_names.values()],
        "calculation": {
            "mode": getattr(workbook.calculation, "calcMode", None),
            "full_calc_on_load": getattr(workbook.calculation, "fullCalcOnLoad", None),
            "force_full_calc": getattr(workbook.calculation, "forceFullCalc", None),
        },
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=cell_value_for_json))
    return 0


def parse_range(worksheet, range_text: str | None):
    if range_text:
        return worksheet[range_text]
    return worksheet[worksheet.calculate_dimension()]


def export_sheet(args) -> int:
    workbook = load_workbook(Path(args.workbook), data_only=args.data_only, read_only=True)
    worksheet = sheet_by_name(workbook, args.sheet)
    rows = []
    for row in parse_range(worksheet, args.range):
        rows.append([cell.value for cell in row])

    if args.format == "json":
        print(json.dumps(rows, ensure_ascii=False, indent=2, default=cell_value_for_json))
    elif args.format == "csv":
        writer = csv.writer(sys.stdout)
        writer.writerows(rows)
    else:
        for row in rows:
            print("\t".join("" if value is None else str(value) for value in row))
    return 0


def find_in_workbook(args) -> int:
    workbook = load_workbook(Path(args.workbook), data_only=args.data_only, read_only=True)
    flags = 0 if args.case_sensitive else re.I
    pattern = re.compile(re.escape(args.pattern), flags)
    matches = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                text = str(value)
                if pattern.search(text):
                    matches.append({"sheet": worksheet.title, "cell": cell.coordinate, "value": text})

    print(json.dumps(matches, ensure_ascii=False, indent=2, default=cell_value_for_json))
    return 0 if matches else 1


def coerce_value(value: str, value_type: str):
    if value_type == "string":
        return value
    if value_type == "number":
        return float(value) if any(mark in value for mark in [".", "e", "E"]) else int(value)
    if value_type == "boolean":
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "y"}:
            return True
        if lowered in {"false", "0", "no", "n"}:
            return False
        raise SystemExit(f"Cannot parse boolean value: {value}")
    if value_type == "blank":
        return None
    if value_type == "formula":
        return value if value.startswith("=") else f"={value}"

    lowered = value.strip().lower()
    if lowered in {"true", "false"}:
        return lowered == "true"
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def save_workbook(workbook, input_path: Path, output_path: Path, allow_overwrite: bool):
    if input_path.resolve() == output_path.resolve() and not allow_overwrite:
        raise SystemExit("Refusing to overwrite the input file without --allow-overwrite.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


def set_cell(args) -> int:
    input_path = Path(args.input_workbook)
    output_path = Path(args.output_workbook)
    workbook = load_workbook(input_path, data_only=False, read_only=False)
    worksheet = sheet_by_name(workbook, args.sheet)
    cell = worksheet[args.cell]
    cell.value = coerce_value(args.value, args.type)
    if args.number_format:
        cell.number_format = args.number_format
    if args.force_recalculate:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    save_workbook(workbook, input_path, output_path, args.allow_overwrite)
    print(json.dumps({"output": str(output_path), "sheet": args.sheet, "cell": args.cell}, indent=2))
    return 0


def replace_text(args) -> int:
    input_path = Path(args.input_workbook)
    output_path = Path(args.output_workbook)
    workbook = load_workbook(input_path, data_only=False, read_only=False)
    flags = 0 if args.case_sensitive else re.I
    pattern = re.compile(re.escape(args.find), flags)
    sheet_names = args.sheets or workbook.sheetnames
    remaining = args.max_count
    count = 0

    for sheet_name in sheet_names:
        worksheet = sheet_by_name(workbook, sheet_name)
        for cell in iter_used_cells(worksheet):
            if remaining is not None and remaining <= 0:
                break
            if isinstance(cell.value, str):
                new_value, changed = pattern.subn(
                    args.replace,
                    cell.value,
                    count=0 if remaining is None else remaining,
                )
                if changed:
                    cell.value = new_value
                    count += changed
                    if remaining is not None:
                        remaining -= changed

    if args.force_recalculate:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    save_workbook(workbook, input_path, output_path, args.allow_overwrite)
    print(json.dumps({"replacements": count, "output": str(output_path)}, indent=2))
    return 0


def csv_to_xlsx(args) -> int:
    openpyxl = require_openpyxl()
    input_path = Path(args.input_csv)
    output_path = Path(args.output_workbook)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = args.sheet_name

    with input_path.open("r", newline="", encoding=args.encoding) as handle:
        reader = csv.reader(handle, delimiter=args.delimiter)
        for row in reader:
            worksheet.append(row)

    if args.freeze_header:
        worksheet.freeze_panes = "A2"
    if args.auto_filter and worksheet.max_row and worksheet.max_column:
        worksheet.auto_filter.ref = worksheet.calculate_dimension()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))
    print(json.dumps({"output": str(output_path), "rows": worksheet.max_row, "columns": worksheet.max_column}, indent=2))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect", help="Inspect workbook structure.")
    inspect_parser.add_argument("workbook")
    inspect_parser.set_defaults(func=inspect_workbook)

    sheet_parser = subparsers.add_parser("sheet", help="Export a sheet range.")
    sheet_parser.add_argument("workbook")
    sheet_parser.add_argument("--sheet", required=True)
    sheet_parser.add_argument("--range")
    sheet_parser.add_argument("--format", choices=["text", "csv", "json"], default="text")
    sheet_parser.add_argument("--data-only", action="store_true", help="Read cached formula results instead of formulas.")
    sheet_parser.set_defaults(func=export_sheet)

    find_parser = subparsers.add_parser("find", help="Find exact text in workbook cells.")
    find_parser.add_argument("workbook")
    find_parser.add_argument("pattern")
    find_parser.add_argument("--case-sensitive", action="store_true")
    find_parser.add_argument("--data-only", action="store_true", help="Search cached formula results instead of formulas.")
    find_parser.set_defaults(func=find_in_workbook)

    set_parser = subparsers.add_parser("set-cell", help="Set one cell value.")
    set_parser.add_argument("input_workbook")
    set_parser.add_argument("output_workbook")
    set_parser.add_argument("--sheet", required=True)
    set_parser.add_argument("--cell", required=True)
    set_parser.add_argument("--value", required=True)
    set_parser.add_argument("--type", choices=["auto", "string", "number", "boolean", "formula", "blank"], default="auto")
    set_parser.add_argument("--number-format")
    set_parser.add_argument("--force-recalculate", action="store_true")
    set_parser.add_argument("--allow-overwrite", action="store_true")
    set_parser.set_defaults(func=set_cell)

    replace_parser = subparsers.add_parser("replace", help="Replace text in string cells.")
    replace_parser.add_argument("input_workbook")
    replace_parser.add_argument("output_workbook")
    replace_parser.add_argument("--find", required=True)
    replace_parser.add_argument("--replace", required=True)
    replace_parser.add_argument("--sheets", nargs="*")
    replace_parser.add_argument("--case-sensitive", action="store_true")
    replace_parser.add_argument("--max-count", type=int)
    replace_parser.add_argument("--force-recalculate", action="store_true")
    replace_parser.add_argument("--allow-overwrite", action="store_true")
    replace_parser.set_defaults(func=replace_text)

    csv_parser = subparsers.add_parser("csv-to-xlsx", help="Convert CSV to XLSX.")
    csv_parser.add_argument("input_csv")
    csv_parser.add_argument("output_workbook")
    csv_parser.add_argument("--sheet-name", default="Data")
    csv_parser.add_argument("--delimiter", default=",")
    csv_parser.add_argument("--encoding", default="utf-8-sig")
    csv_parser.add_argument("--freeze-header", action="store_true")
    csv_parser.add_argument("--auto-filter", action="store_true")
    csv_parser.set_defaults(func=csv_to_xlsx)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
